"""Load the underwriting scoring framework from the team's shared Excel file
(config.settings.SCORING_FRAMEWORK_PATH, currently Underwriting_POC.xlsx).

This is the single source of truth for scoring rules — factor bands, weights,
and risk-band thresholds/actions live in that spreadsheet, owned by the
domain expert. DO NOT hand-copy scores into Python; read them fresh here so
an updated spreadsheet takes effect on the next pipeline run automatically.

Expected sheets:
  "Framework" — columns: Factor | Condition/band | Raw score | Weight |
                Weighted contribution | Interpretation
                4 rows per factor (raw scores 0, 1, 2, 3); weight is constant
                across a factor's 4 rows.
  "Bands"     — columns: Min total score | Max total score | Risk band |
                Suggested underwriting action
                Max total score may be an open-ended string (e.g. "37+") for
                the top band.
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config import settings  # noqa: E402


def load_framework(path: Path | None = None) -> list[dict]:
    """Return one entry per (factor, band) row from the Framework sheet:
    {factor, band_label, raw_score, weight, weighted_contribution, interpretation}
    """
    import openpyxl

    path = path or settings.SCORING_FRAMEWORK_PATH
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Framework"]

    rows = list(ws.iter_rows(values_only=True))
    entries: list[dict] = []
    for row in rows[1:]:  # skip header
        if not row or row[0] is None:
            continue
        factor, band_label, raw_score, weight, weighted_contribution, interpretation = row[:6]
        entries.append(
            {
                "factor": factor.strip() if isinstance(factor, str) else factor,
                "band_label": band_label.strip() if isinstance(band_label, str) else band_label,
                "raw_score": int(raw_score),
                "weight": int(weight),
                "weighted_contribution": int(weighted_contribution),
                "interpretation": interpretation,
            }
        )
    return entries


def load_bands(path: Path | None = None) -> list[dict]:
    """Return the risk-band lookup table:
    {min_score, max_score, band, suggested_action}
    max_score is None for an open-ended top band (e.g. "37+").
    """
    import openpyxl

    path = path or settings.SCORING_FRAMEWORK_PATH
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Bands"]

    rows = list(ws.iter_rows(values_only=True))
    bands: list[dict] = []
    for row in rows[1:]:  # skip header
        if not row or row[0] is None:
            continue
        min_score, max_score, band, action = row[:4]
        if isinstance(max_score, str):  # e.g. "37+" -> open-ended
            max_score = None
        bands.append(
            {
                "min_score": int(min_score),
                "max_score": int(max_score) if max_score is not None else None,
                "band": band,
                "suggested_action": action,
            }
        )
    return bands


def factors_grouped(framework_entries: list[dict]) -> dict[str, list[dict]]:
    """Group framework rows by factor name, preserving sheet order (which is
    also raw_score-ascending: 0, 1, 2, 3)."""
    grouped: dict[str, list[dict]] = {}
    for entry in framework_entries:
        grouped.setdefault(entry["factor"], []).append(entry)
    return grouped


def lookup_band(total_score: int, bands: list[dict]) -> dict:
    """Find the band whose [min_score, max_score] range contains total_score.
    The top band has max_score=None (open-ended, e.g. "37+")."""
    for band in bands:
        if band["max_score"] is None:
            if total_score >= band["min_score"]:
                return band
        elif band["min_score"] <= total_score <= band["max_score"]:
            return band
    raise ValueError(
        f"No band found for total_score={total_score} — check the Bands sheet "
        "covers this range."
    )
